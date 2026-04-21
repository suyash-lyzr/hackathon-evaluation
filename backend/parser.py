"""Excel submission parsing."""
from io import BytesIO
from openpyxl import load_workbook


COLUMN_ALIASES = {
    "team_name":       ["team name", "team", "team_name"],
    "project_title":   ["project title", "project", "title", "project_title"],
    "elevator_pitch":  ["short elevator pitch (3-4 liners)", "elevator pitch", "pitch", "short elevator pitch", "elevator_pitch"],
    "live_url":        ["live deployed url", "live url", "deployed url", "url", "live_url"],
    "app_id":          ["app id", "app_id", "appid", "app id:", "architect app id"],
    "pain_point":      ["what specific \"pain point\" does this solve?", "pain point", "what specific 'pain point' does this solve?", "pain_point", "what pain point does this solve"],
    "primary_user":    ["who is the primary user?", "primary user", "primary_user", "target user", "user"],
    "impact":          ["quantify the impact", "impact", "business impact", "impact:", "quantify the impact:"],
    "loom_video":      ["loom video", "loom", "loom_video", "video"],
}


def _canonicalize_header(raw: str) -> str | None:
    if not raw:
        return None
    key = str(raw).strip().lower().rstrip(":").strip()
    for canon, aliases in COLUMN_ALIASES.items():
        if key in [a.lower() for a in aliases]:
            return canon
        if key == canon:
            return canon
    return None


def parse_submissions_xlsx(file_bytes: bytes) -> list[dict]:
    """Parse an uploaded xlsx. Returns list of row dicts keyed by our canonical fields.

    Rows missing an app_id are kept but flagged — the UI will surface them.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Header row = first non-empty row
    header_idx = 0
    for idx, r in enumerate(rows):
        if any(c is not None and str(c).strip() for c in r):
            header_idx = idx
            break

    header = rows[header_idx]
    col_map: dict[int, str] = {}
    for i, cell in enumerate(header):
        canon = _canonicalize_header(cell)
        if canon:
            col_map[i] = canon

    submissions: list[dict] = []
    for r in rows[header_idx + 1 :]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        entry: dict[str, str] = {k: "" for k in COLUMN_ALIASES}
        for i, cell in enumerate(r):
            canon = col_map.get(i)
            if canon and cell is not None:
                entry[canon] = str(cell).strip()
        if not any(entry.values()):
            continue
        submissions.append(entry)

    return submissions
