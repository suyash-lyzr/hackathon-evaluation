"""Generates a sample submissions xlsx with dummy data + one real app_id for testing."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

OUT = Path(__file__).resolve().parent.parent / "sample_submissions.xlsx"

HEADERS = [
    "Team Name",
    "Project Title",
    "Short Elevator Pitch (3-4 liners)",
    "Live Deployed URL",
    "App ID",
    "What specific \"pain point\" does this solve?",
    "Who is the primary user?",
    "Quantify the Impact",
    "Loom Video",
]

ROWS = [
    [
        "Zenith Labs",
        "HireOps — Autonomous Recruiter",
        "Multi-agent recruiter that sources, screens, and schedules interviews. One orchestrator routes candidates across sourcing, screening, and coordination agents. Built for small HR teams that can't afford an applicant tracking stack.",
        "https://hireops.architect.new",
        "69e7161fff1ef4a43123e111",
        "Small companies lose 4-6 hours per day on candidate sourcing, screening calls, and scheduling. Recruiters waste time on low-signal resumes and back-and-forth calendaring instead of talking to real candidates.",
        "In-house recruiters and hiring managers at 10-200 person companies without a dedicated TA function.",
        "Saves 20 hrs/week per recruiter. Replaces ~$400/mo in tooling (LinkedIn Recruiter Lite + Calendly + manual Google Sheets).",
        "https://loom.com/share/example-hireops",
    ],
    [
        "Finlytic",
        "Invoice Reconciler",
        "An agent that reads incoming PDF invoices, matches them to POs, flags mismatches, and drafts approval emails. Runs nightly; humans review only exceptions.",
        "https://finlytic.architect.new",
        "000000000000000000000000",
        "AP teams spend 30% of their week manually reconciling invoices to purchase orders. Errors cost an average mid-market company ~$50k/yr in duplicate or late payments.",
        "Accounts payable analysts and finance operations leads at mid-market B2B SaaS and services firms.",
        "Cuts manual reconciliation time by 70%. For a 3-person AP team that's ~60 hours/week saved.",
        "",
    ],
    [
        "NeighborNet",
        "Community Event Planner",
        "Helps neighborhood associations plan events. Single agent guides the user through checklist questions, then outputs a plan.",
        "https://neighbornet.architect.new",
        "000000000000000000000001",
        "People who organize local events don't know where to start.",
        "Anyone.",
        "Saves time.",
        "",
    ],
    [
        "PulseCare",
        "Clinic Intake & Triage",
        "Patient intake agent collects symptoms, pulls history via EHR integration, and routes to nurse queue with urgency score. Coordinator agent loops in an on-call physician for high-urgency cases.",
        "https://pulsecare.architect.new",
        "ffffffffffffffffffffffff",
        "Primary care clinics lose 15+ min per patient on intake. Triage happens at the front desk with no urgency signal, so high-risk patients sit in the waiting room next to routine checkups.",
        "Primary care front-desk staff and nurse triage leads in independent/family clinics (1-10 provider offices).",
        "Saves 12 min per patient intake. At 40 patients/day that's 8 hrs/day reclaimed per clinic.",
        "https://loom.com/share/example-pulsecare",
    ],
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6B4C4C", end_color="6B4C4C", fill_type="solid")

    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for r, row in enumerate(ROWS, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [18, 28, 60, 36, 30, 60, 40, 50, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 36
    for r in range(2, len(ROWS) + 2):
        ws.row_dimensions[r].height = 80

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
